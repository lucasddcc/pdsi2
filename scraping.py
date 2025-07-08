import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, Table, Column, Integer, Text, TIMESTAMP, MetaData
from dotenv import load_dotenv
import logging
from datetime import datetime
import os

# Configuração
load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Estilos Excel
orangeFill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
blueFill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')

# Banco de Dados
DB_URL = f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
engine = create_engine(DB_URL)
metadata = MetaData()

menu_table = Table('menu_ufu', metadata,
    Column('id', Integer, primary_key=True),
    Column('item_menu', Text),
    Column('link', Text),
    Column('data_coleta', TIMESTAMP)
)

# Web Scraping
def scrape_ufu_menu():
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    ws = wb.create_sheet('planilha')
    
    # Cabeçalhos
    ws.cell(1, 1, "Menu Nav").fill = orangeFill
    ws.cell(1, 2, "Links").fill = blueFill

    try:
        response = requests.get('https://ufu.br/', timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        
        leftBar = soup.find('ul', class_='sidebar-nav nav-level-0')
        leftBarLines = leftBar.find_all('li', class_='nav-item') if leftBar else []
        
        init_capture = False
        row = 2  # Começa na linha 2 (abaixo do cabeçalho)
        
        with engine.connect() as conn:
            conn.execute(menu_table.delete())  # Limpa tabela existente
            
            for li in leftBarLines:
                text = li.get_text(strip=True)
                link = li.a.get('href') if li.a else ''
                
                # Ativa captura após encontrar "Graduação"
                if 'Graduação' in text:
                    init_capture = True
                
                if init_capture and text and link:
                    full_link = f"https://ufu.br{link}" if link.startswith('/') else link
                    
                    # Excel
                    ws.cell(row, 1, text)
                    ws.cell(row, 2, full_link)
                    
                    # PostgreSQL
                    conn.execute(
                        menu_table.insert().values(
                            item_menu=text,
                            link=full_link,
                            data_coleta=datetime.now()
                        )
                    )
                    row += 1
            
            conn.commit()
            logger.info(f"{row-2} itens salvos no banco de dados")
        
        wb.save("UFU_menu_nav.xlsx")
        logger.info("Planilha gerada: UFU_menu_nav.xlsx")
        
    except Exception as e:
        logger.error(f"Erro: {str(e)}")
    finally:
        engine.dispose()

if __name__ == "__main__":
    scrape_ufu_menu()